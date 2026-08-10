#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""服饰创意分析站：总站标签树 + 三个行业深度页。

- 总站仅呈现行业整体标签树和四个赛道入口。
- 箱包鞋靴入口直达既有深度站，内容保持原样。
- 内衣、男装、女装分别生成与箱包鞋靴一致的三 Tab 页面：
  标签树 / 标签分析和结论 / 素材内容逐帧剖析。
"""
from __future__ import annotations

import hashlib
import html
import json
import shutil
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent
XLSX = Path('/Users/chenxiaomei/Desktop/服饰大盘.xlsx')
DOCS = BASE / 'docs'
DATA_DIR = BASE / 'data'
FRAME_ROOT = BASE / 'assets' / 'frames'

SECTORS = {
    'underwear': {'name': '内衣', 'emoji': '🩲', 'sheet': '内衣', 'desc': '舒适、支撑与贴身功能表达'},
    'menswear': {'name': '男装', 'emoji': '👔', 'sheet': '男装', 'desc': '功能、版型与通勤场景表达'},
    'womenswear': {'name': '女装', 'emoji': '👗', 'sheet': '女装', 'desc': '上身效果、风格与穿搭种草'},
}

# 分行业/类目创意标签基线。实际关键帧视觉覆盖优先写在 manual_visual_tags.json。
CATEGORY_TAGS = {
    '文胸/乳贴/内裤': {'struct':'真人口播','role':'单人女','scene':'居家室内','sell':'功能实测','function':'舒适贴合','emotion':'身材焦虑自信种草','people':'女性贴身需求','pain':'勒痕/空杯/闷热'},
    '袜子': {'struct':'沉浸式产品展示','role':'无人出镜','scene':'居家室内','sell':'功能实测','function':'吸汗透气','emotion':'促销紧迫感','people':'日常通勤','pain':'闷脚/滑落/磨脚'},
    '塑身衣/裤': {'struct':'真人口播','role':'单人女','scene':'居家室内','sell':'场景痛点','function':'收腹提臀','emotion':'身材焦虑自信种草','people':'女性塑形','pain':'显胖/勒肉/卷边'},
    '睡衣/家居服': {'struct':'真人口播','role':'单人女','scene':'居家室内','sell':'颜值种草','function':'亲肤舒适','emotion':'舒适治愈','people':'居家女性','pain':'闷热/扎肤/版型松垮'},
    '儿童内衣裤袜': {'struct':'真人口播','role':'单人女','scene':'居家室内','sell':'信任背书','function':'亲肤安全','emotion':'家长关怀','people':'儿童家长','pain':'敏感肌/勒痕/不耐穿'},
    '上衣': {'struct':'真人口播','role':'单人女','scene':'居家室内','sell':'颜值种草','function':'显瘦修饰','emotion':'身材焦虑自信种草','people':'通勤女性','pain':'显胖/不好搭/版型不合身'},
    '裤子': {'struct':'真人口播','role':'单人女','scene':'居家室内','sell':'功能实测','function':'显腿长','emotion':'身材焦虑自信种草','people':'日常通勤','pain':'卡裆/显腿粗/不舒服'},
    '裙子': {'struct':'真人口播','role':'单人女','scene':'户外街景','sell':'颜值种草','function':'显瘦显高','emotion':'美女吸睛','people':'约会穿搭','pain':'显胖/透/不好搭'},
    '套装/学生校服/工作制服': {'struct':'真人口播','role':'单人女','scene':'居家室内','sell':'场景痛点','function':'省心搭配','emotion':'省心决策','people':'通勤/学生','pain':'搭配困难/不正式'},
    '速干衣裤': {'struct':'真人口播','role':'单人男','scene':'户外街景','sell':'功能实测','function':'速干透气','emotion':'运动挑战','people':'运动户外','pain':'闷汗/黏身/不耐磨'},
    '防晒衣/皮肤衣': {'struct':'真人口播','role':'单人女','scene':'户外街景','sell':'功能实测','function':'防晒透气','emotion':'夏日出行','people':'户外女性','pain':'晒黑/闷热/不显瘦'},
    '羽绒服/棉服': {'struct':'真人口播','role':'单人女','scene':'户外街景','sell':'功能实测','function':'保暖轻量','emotion':'冬季保暖','people':'冬季通勤','pain':'臃肿/不保暖/钻绒'},
    '衬衫': {'struct':'真人口播','role':'单人女','scene':'居家室内','sell':'颜值种草','function':'挺括显瘦','emotion':'通勤自信','people':'职场通勤','pain':'易皱/显胖/不好搭'},
}
DEFAULT_TAGS = {'struct':'真人口播','role':'单人女','scene':'居家室内','sell':'功能实测','function':'核心卖点展示','emotion':'结果导向','people':'目标消费人群','pain':'常见决策顾虑'}

# 为分析页补齐与箱包鞋靴一致的全部标签维度；基于各赛道品类和素材主题的基线，关键帧人工标签仍优先覆盖结构/角色/场景。
SECTOR_EXTRA_TAGS = {
    'underwear': {'season':'四季通用','material':'纯棉/莫代尔','style':'简约舒适','use':'日常贴身','festival':'无节点'},
    'menswear': {'season':'四季通勤','material':'纯棉/聚酯纤维','style':'商务简约','use':'商务通勤','festival':'无节点'},
    'womenswear': {'season':'夏季','material':'针织/雪纺','style':'通勤简约','use':'日常通勤','festival':'换季'},
}

# 服饰行业整体标签树：复刻用户给出的完整标签框架，作为总站的统一口径。
# 字段：一级标签、一级说明、是否核心、[(二级标签、优先级、Value枚举、判断依据、一级分类关联说明、二级分类说明、开屏/种草)]
OVERALL_TREE = [
    ('商品属性', '商品本身的可见属性与购买基础决策点。', False, [
        ('适用季节', 'p0', '春季 / 夏季 / 秋季 / 冬季 / 四季通用 / 开学季 / 暑假出游季 / 换季', '商品厚薄、画面环境、字幕或口播中的季节词。', '支撑季节性投放与节点规划。', '商品适用季节、支撑季节性投放节奏。', '开屏通过季节词或场景快速建立需求。'),
        ('功能属性', 'p0', '防滑 / 防水防泼 / 透气 / 显瘦显高 / 增高 / 大容量 / 轻便折叠 / 收纳分层 / 耐磨 / 缓震回弹 / 护脊减负', '文字/字幕功能词为主，画面实测动作辅助确认。', '决定商品主打卖点与转化核心。', '商品功能属性、服务细分需求。', '素材主打的卖点元素，创建可感知结果。'),
        ('材质属性', 'p0', '纯棉 / 头层牛皮 / PU仿皮 / 冰丝 / 莫代尔 / 帆布织物 / 橡胶大底 / EVA / 尼龙防泼水', '画面材质特写 + 字幕材质词；无清晰依据不强行标注。', '建立材质信任，影响高客单决策。', '画面中被强调的材质信息。', '常作卖点，强化质感与可信度。'),
    ]),
    ('内容形式', '最核心：先判断素材如何被拍成一个可消费的视频。', True, [
        ('制作形式', 'p00', '真人口播 / 多人剧情 / 沉浸式产品展示 / 直播切片 / AI数字人口播 / 图文', '有无真人对镜头说话；是否多人互动/剧情冲突；是否仅产品/手部；是否有直播浮层、AI合成痕迹或静态图文排版。', '素材创意体裁的第一切分维度。', '素材制作形式，效果归因第一切分维度。', '前3秒的制作形式决定是否进入观看与后续转化。'),
    ]),
    ('卖点表达', '素材如何组织卖点、降低理解成本并建立信任。', False, [
        ('功能实测', 'p0', '弹力拉伸 / 防水防泼 / 透气 / 耐磨 / 承重 / 收纳 / 防滑 / 减震 / 回弹', '口播/字幕关键词 + 测试动作综合判断。', '将功能承诺转成可见证据。', '测试动作与对应功能词形成闭环。', '素材主打的卖点元素，创建可感知结果。'),
        ('场景痛点', 'p0', '磨脚 / 勒肉 / 闷热 / 显胖 / 显矮 / 容量不足 / 收纳凌乱 / 易皱 / 晒黑', '痛点文字、用户口播、前后对比或场景冲突。', '从用户问题切入，提升代入感。', '描述目标用户的真实痛点与场景。', '切入用户痛点，影响共鸣与转化。'),
        ('信任背书类型', 'p0', '品牌历史 / 工艺 / 产地溯源 / 检测报告 / 达人推荐 / 长辈祝福', '字幕/口播中的背书信息，以及工厂、报告、专业身份画面。', '建立可信度，影响高客单转化。', '增强可信度的内容表达。', '增加可信度元素，影响高客单转化。'),
        ('价格/促销表达', 'p0', '买一送一 / 限时折扣 / 满减 / 包邮到家 / 工厂直销 / 直播专属价', '价格角标、口播促销词、直播价签。', '促销信息影响即时转化。', '价格与促销的具体表达。', '价格与促销信息呈现，影响CVR。'),
        ('竞品对比', 'p0', '材质对比 / 工艺对比 / 版型对比 / 鞋底对比 / 容量对比 / 价格对比', '同屏或前后对照画面、明确对比话术。', '强化差异化购买理由。', '通过对比突出产品优势。', '开屏或中段用对比快速建立优势。'),
    ]),
    ('情绪与营销', '通过情绪、场景、节点与风格建立观看动机和人群代入。', False, [
        ('情绪钩子类型', 'p0', '身材焦虑 / 大牌平替 / 促销紧迫感 / 美女吸睛 / 亲情关怀 / 猎奇实验 / 舒适治愈', '开头3秒标题、口播话术和画面情绪。', '决定点击与短期留存。', '情绪与人群需求的创意表达。', '情绪钩子决定 3 秒完播与点击。'),
        ('题材/使用场景', 'p0', '日常通勤 / 旅行出行 / 运动户外 / 居家生活 / 礼赠 / 中老年关怀 / 工厂溯源', '场景背景 + 口播用途综合判定。', '使商品进入真实生活决策。', '素材服务的生活用途与场景。', '支持场景化叙事与人群定向。'),
        ('节日/节点营销', 'p0', '开学季 / 520 / 儿童节 / 618 / 双11 / 春节 / 暑假出游 / 无节点', '字幕节点词、促销话术和节日画面。', '判断节点内容的扩量机会。', '营销节点及相应表达。', '节点词可直接进入开场或权益收尾。'),
        ('穿搭风格类型', 'p0', '通勤简约 / 休闲日常 / 御姐辣妹 / 商务正装 / 街头潮酷 / 轻奢高级 / 少女法系 / 运动风', '整体视觉调性、版型、配色综合判断。', '衡量审美与目标人群匹配度。', '风格属性，支撑风格/人群/效果关联。', '风格决定用户第一印象和穿搭关系。'),
    ]),
    ('画面客观要素', '最核心：根据真实画面识别拍摄地点与人物构成。', True, [
        ('拍摄场景类型', 'p00', '居家室内 / 纯色棚拍 / 酒店大堂 / 户外街景 / 工厂车间 / 机场 / 商场门店 / 学校展会', '画面物理拍摄地直接识别。', '真实场景、棚拍和专业场景影响观看与转化。', '拍摄画面环境，影响场景真实感与代入。', '拍摄背景决定内容可信度和沉浸感。'),
        ('出镜角色类型', 'p00', '单人女 / 单人男 / 多人对话 / 无人出镜（纯产品/手部） / 明星', '画面人物数量、性别、关系识别；仅手部/纯产品归无人出镜。', '人物出镜决定人群感知与代入。', '出镜角色类型，影响内容人设与代入。', '角色是最关键的用户关系建立要素。'),
    ]),
]
# 所有行业标签树都遵从核心优先顺序：内容形式、画面客观要素在前。
TREE_ORDER = {'内容形式': 0, '画面客观要素': 1, '商品属性': 2, '卖点表达': 3, '情绪与营销': 4}
OVERALL_TREE.sort(key=lambda item: TREE_ORDER[item[0]])

SECTOR_VALUES = {
    'underwear': {
        '商品属性': [('适用季节','p0','四季通用 / 夏季凉感 / 秋冬保暖 / 孕产 / 少女期','厚薄、凉感/保暖面料和文案季节词。'),('功能属性','p0','无痕 / 舒适贴合 / 支撑聚拢 / 收腹提臀 / 透气抗菌 / 吸汗 / 保暖 / 防磨腿','贴身试穿、弹力/面料特写及字幕。'),('材质属性','p0','纯棉 / 莫代尔 / 冰丝 / 蕾丝 / 桑蚕丝 / 乳胶 / 锦纶氨纶','材质特写、吊牌或口播材质词。')],
        '卖点表达': [('功能实测','p0','弹力拉伸 / 无痕对比 / 支撑展示 / 吸汗透气 / 防卷边','试穿、拉伸、局部特写和字幕。'),('场景痛点','p0','空杯 / 勒痕 / 副乳 / 闷热 / 内裤线 / 磨腿 / 卷边','痛点口播、前后对比或穿着局部。'),('信任背书/促销/对比','p0','面料安全 / 工厂 / 专利 / 多件装 / 限时优惠 / 旧款对比','报告/工厂画面、价格浮层和对比话术。')],
        '情绪与营销': [('情绪钩子','p0','身材自信 / 舒适治愈 / 家长关怀 / 促销紧迫感','开头口播、字幕和情绪表达。'),('使用场景/人群','p0','日常贴身 / 居家睡眠 / 运动 / 孕产 / 少女 / 儿童','背景、人物及用途口播。'),('风格/节点','p0','简约舒适 / 甜美蕾丝 / 轻塑形 / 开学季 / 618 / 无节点','视觉调性和节点字幕。')],
    },
    'menswear': {
        '商品属性': [('适用季节','p0','春夏薄款 / 夏季速干 / 秋冬保暖 / 四季通勤','商品厚薄、场景和季节字幕。'),('功能属性','p0','速干透气 / 凉感 / 弹力 / 抗皱 / 耐磨 / 保暖 / 宽松显瘦','面料/动作实测和字幕功能词。'),('材质属性','p0','纯棉 / 亚麻 / 冰丝 / 聚酯纤维 / 牛仔 / 羊毛 / 锦纶','面料特写与口播材质词。')],
        '卖点表达': [('功能实测','p0','速干实验 / 透气测试 / 拉伸 / 抗皱 / 耐磨 / 防晒','测试动作与结果字幕。'),('场景痛点','p0','闷汗 / 易皱 / 显肚 / 卡裆 / 不耐穿 / 通勤不正式','男性真实口播、前后对比或通勤场景。'),('信任背书/促销/对比','p0','工厂直供 / 面料工艺 / 大牌平替 / 多件套 / 限时价','工厂、口播身份、价格与对比信息。')],
        '情绪与营销': [('情绪钩子','p0','实用省心 / 爸爸关怀 / 兄弟推荐 / 平替心理 / 运动挑战','开头标题与人设口播。'),('使用场景/人群','p0','商务通勤 / 户外运动 / 父辈 / 旅行 / 学生','人物人设、背景和用途口播。'),('风格/节点','p0','商务简约 / 休闲日常 / 工装户外 / 运动风 / 父亲节 / 换季','版型、颜色和节点字幕。')],
    },
    'womenswear': {
        '商品属性': [('适用季节','p0','春季 / 夏季 / 秋季 / 冬季 / 四季通勤 / 换季','商品厚薄、背景季节与口播。'),('功能属性','p0','显瘦 / 显高 / 遮胯 / 修饰肩颈 / 垂坠 / 防晒 / 凉感 / 弹力','上身对比、走动动态与字幕功能词。'),('材质属性','p0','棉 / 麻 / 真丝 / 醋酸 / 针织 / 雪纺 / 牛仔 / 蕾丝','面料近景、手感和材质字幕。')],
        '卖点表达': [('功能实测','p0','上身前后对比 / 多身材试穿 / 拉伸 / 垂坠 / 防透展示','真人试穿、局部近景和结果字幕。'),('场景痛点','p0','显胖 / 显矮 / 不好搭 / 透 / 卡肩 / 遮不住胯','身材焦虑口播、前后对比和搭配演示。'),('信任背书/促销/对比','p0','工厂源头 / 面料工艺 / 多色多码 / 限时优惠 / 大牌平替','工厂/面料画面、价格浮层和对比话术。')],
        '情绪与营销': [('情绪钩子','p0','身材自信 / 美女吸睛 / 约会氛围 / 平替心理 / 换季焕新','开头标题、人物状态和口播。'),('使用场景/人群','p0','通勤 / 约会 / 旅行 / 母女 / 微胖 / 小个子','人物、场景与用途口播。'),('风格/节点','p0','通勤简约 / 少女法系 / 御姐辣妹 / 新中式 / 轻奢 / 夏日度假 / 618','整体造型、配色和节点字幕。')],
    },
}


def num(v):
    try: return float(v)
    except (TypeError, ValueError): return 0.0


def clean(v):
    v = '' if v is None else str(v).strip()
    return '' if v in {'空','None','nan','-'} else v


def md5(url): return hashlib.md5(url.encode()).hexdigest()[:10]


def load_overrides():
    p = DATA_DIR / 'manual_visual_tags.json'
    return json.loads(p.read_text(encoding='utf-8')) if p.exists() else {}


def load_sector(sheet, sector):
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[sheet]
    overrides = load_overrides().get(sector, {})
    rows = {}
    for r in ws.iter_rows(min_row=2, max_col=8, values_only=True):
        _, url, cat2, cat3, spend, ctr, cvr, v3 = r
        if not isinstance(url, str) or not url.startswith('http'): continue
        row = {'id':md5(url),'url':url,'cat2':clean(cat2),'category':clean(cat3) or clean(cat2),'spend':num(spend),'ctr':num(ctr),'cvr':num(cvr),'v3':num(v3)}
        if url not in rows or row['spend'] > rows[url]['spend']: rows[url] = row
    out = list(rows.values())
    for r in out:
        tags = dict(SECTOR_EXTRA_TAGS[sector])
        tags.update(CATEGORY_TAGS.get(r['category'], DEFAULT_TAGS))
        tags.update(overrides.get(r['id'], {}))
        r['tags'] = tags
    return sorted(out, key=lambda x:x['spend'], reverse=True)


def wavg(rows, key):
    total = sum(x['spend'] for x in rows)
    return sum(x['spend']*x[key] for x in rows)/total if total else 0


def aggregate(rows, tag):
    total = sum(r['spend'] for r in rows)
    groups = defaultdict(list)
    for r in rows: groups[r['tags'][tag]].append(r)
    result=[]
    for name, rs in groups.items():
        cost=sum(x['spend'] for x in rs)
        result.append({'name':name,'count':len(rs),'share':round(cost/total*100,1) if total else 0,'ctr':round(wavg(rs,'ctr'),2),'cvr':round(wavg(rs,'cvr'),2),'v3':round(wavg(rs,'v3'),1)})
    # 结构类型固定补齐五类
    if tag == 'struct':
        for name in ['真人口播','多人剧情','沉浸式产品展示','直播切片','AI数字人口播']:
            if name not in groups: result.append({'name':name,'count':0,'share':None,'ctr':None,'cvr':None,'v3':None})
    return sorted(result,key=lambda x:(x['count']==0,-(x['share'] or 0)))


def categories(rows):
    d=defaultdict(list)
    for r in rows:
        if r['category']: d[r['category']].append(r)
    total=sum(r['spend'] for r in rows)
    out=[]
    for name,rs in d.items():
        cost=sum(r['spend'] for r in rs)
        out.append({'name':name,'count':len(rs),'share':round(cost/total*100,1),'ctr':round(wavg(rs,'ctr'),2),'cvr':round(wavg(rs,'cvr'),2),'v3':round(wavg(rs,'v3'),1)})
    return sorted(out,key=lambda x:-x['share'])[:10]


def sector_payload(key):
    meta=SECTORS[key]; rows=load_sector(meta['sheet'],key)
    total=sum(r['spend'] for r in rows)
    by={x:aggregate(rows,x) for x in ['struct','role','scene','style','sell','function','material','season','emotion','use','festival']}
    cats=categories(rows)
    active=lambda xs:[x for x in xs if x['count']]
    def leader(xs): return active(xs)[0] if active(xs) else {'name':'—','share':0}
    def best(xs,field): return max(active(xs),key=lambda x:x[field]) if active(xs) else {'name':'—',field:0}
    bs=best(by['struct'],'ctr'); bv=best(by['struct'],'v3')
    narrative=(f"{meta['name']}赛道消耗主要集中在「{leader(cats)['name'] if cats else '—'}」；创意结构以「{leader(by['struct'])['name']}」为主。"
               f"CTR 最优为「{bs['name']}」（{bs['ctr']}%），3秒完播最优为「{bv['name']}」（{bv['v3']}%）。")
    return {'id':key,'name':meta['name'],'emoji':meta['emoji'],'desc':meta['desc'],'metrics':{'materials':len(rows),'ctr':round(wavg(rows,'ctr'),2),'cvr':round(wavg(rows,'cvr'),2),'v3':round(wavg(rows,'v3'),1)},'analysis':by,'categories':cats,'top':rows[:12],'narrative':narrative,'note':'标签先以品类创意基线生成；Top素材的可见画面标签会由 manual_visual_tags.json 的关键帧人工核验覆盖。'}


def esc(v): return html.escape(str(v))


def sector_tree(sector_id: str):
    """各赛道共享核心视觉标签，但商品属性/卖点/营销枚举按行业实际调整。"""
    values = SECTOR_VALUES[sector_id]
    out = []
    for title, desc, core, children in OVERALL_TREE:
        if title in values:
            local_children = [(n, p, vals, rule, '', '', '') for n, p, vals, rule in values[title]]
        else:
            local_children = children
        out.append((title, desc, core, local_children))
    return out


def tree_html(tree, overall=False):
    parts=[]
    for title,desc,core,children in tree:
        if overall:
            cards=''.join(
                f'<tr><td class="child-name">{esc(n)}</td><td><em>{p}</em></td><td>{esc(vals)}</td><td>{esc(rule)}</td><td>{esc(level1)}</td><td>{esc(level2)}</td><td>{esc(opening)}</td></tr>'
                for n,p,vals,rule,level1,level2,opening in children
            )
            body=f'<div class="tree-scroll"><table class="overall-table"><thead><tr><th>二级标签</th><th>优先级</th><th>Value/枚举值</th><th>判断依据 / 画面特征</th><th>一级标签分类说明</th><th>二级标签分类说明</th><th>标签关键点</th></tr></thead><tbody>{cards}</tbody></table></div>'
        else:
            cards=''.join(f'<div class="child"><b>{esc(n)}</b><em>{p}</em><span><strong>标签值：</strong>{esc(vals)}</span><span><strong>判断：</strong>{esc(rule)}</span></div>' for n,p,vals,rule,*_ in children)
            body=f'<div class="tree-child">{cards}</div>'
        parts.append(f'<article class="tree"><div class="tree-main {"core" if core else ""}"><strong>{esc(title)}</strong><span>{esc(desc)}</span></div>{body}</article>')
    return ''.join(parts)


def home_html():
    entries = '<a class="entry" href="https://chenxiaomei19960125-glitch.github.io/bag-shoes-creative-report/"><div class="emoji">👜</div><b>箱包鞋靴</b><span>进入既有深度站：标签树、标签分析与结论、素材逐帧剖析均保持原样。</span></a>'
    for key,m in SECTORS.items():
        entries += f'<a class="entry" href="{key}/"><div class="emoji">{m["emoji"]}</div><b>{m["name"]}</b><span>{m["desc"]}；进入独立深度分析页。</span></a>'
    return f'''<!doctype html><html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>服饰行业创意分析</title>{STYLE}</head><body><main class="wrap"><header class="hero"><div class="eyebrow">AI 创意洞察 · 服饰行业专项</div><h1>服饰行业创意分析</h1><p>总站仅提供服饰行业整体标签树和四个赛道入口；各赛道进入后使用独立的「标签树结构 / 标签分析和结论 / 素材内容逐帧剖析」页面。</p></header><section class="section"><h2>进入赛道分析</h2><p>箱包鞋靴保留已上线的原站内容；内衣、男装和女装分别进入独立分析站。</p><div class="entry-grid">{entries}</div></section><section class="section"><h2>服饰行业整体标签树结构</h2><p>内容形式、画面客观要素为最优先的创意标签；其余标签用于补足商品决策、卖点表达和情绪营销解释。</p><div class="tree-wrap">{tree_html(OVERALL_TREE, overall=True)}</div></section><footer class="footer">服饰创意 XAI · 统一标签语言、效果归因与素材复用</footer></main></body></html>'''

STYLE = r'''<style>
:root{--blue:#0052d9;--ink:#172b4d;--text:#33425c;--muted:#71809a;--bg:#f4f7fb;--card:#fff;--line:#e2eaf5}*{box-sizing:border-box}body{margin:0;background:var(--bg);color:var(--ink);font-family:-apple-system,BlinkMacSystemFont,"PingFang SC","Microsoft YaHei",Arial,sans-serif}.wrap{max-width:1440px;margin:auto;padding:28px 36px 56px}.hero{padding:42px;border-radius:24px;background:linear-gradient(135deg,#003ea6,#0052d9 55%,#1590f4);color:#fff;box-shadow:0 18px 42px rgba(0,82,217,.22);position:relative;overflow:hidden}.hero:after{content:"";position:absolute;right:-80px;top:-130px;width:420px;height:420px;border-radius:50%;background:rgba(255,255,255,.12)}.eyebrow{position:relative;z-index:1;display:inline-block;padding:5px 10px;border:1px solid rgba(255,255,255,.42);border-radius:999px;background:rgba(255,255,255,.13);font-size:12px;font-weight:600}.hero h1{position:relative;z-index:1;margin:16px 0 10px;font-size:34px}.hero p{position:relative;z-index:1;max-width:850px;margin:0;color:rgba(255,255,255,.9);font-size:14px;line-height:1.8}.section{margin-top:20px;padding:28px;background:var(--card);border:1px solid var(--line);border-radius:20px;box-shadow:0 5px 18px rgba(0,82,217,.04)}.section h2{margin:0;font-size:22px}.section>p{color:var(--muted);font-size:13px;line-height:1.75}.entry-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-top:18px}.entry{display:block;text-decoration:none;color:var(--ink);padding:22px;border:1px solid var(--line);border-radius:16px;background:#fff;transition:.2s;box-shadow:0 3px 12px rgba(0,82,217,.04)}.entry:hover{transform:translateY(-3px);border-color:#76aaff;box-shadow:0 12px 25px rgba(0,82,217,.14)}.entry .emoji{font-size:28px}.entry b{display:block;margin:11px 0 5px;font-size:18px}.entry span{font-size:12px;line-height:1.55;color:var(--muted)}.tree-wrap{display:flex;flex-direction:column;gap:14px;margin-top:18px}.tree{display:grid;grid-template-columns:220px minmax(0,1fr);border:1px solid var(--line);border-radius:15px;overflow:hidden;background:#fff}.tree-main{padding:20px;background:linear-gradient(160deg,#eaf2ff,#dceaff);border-right:1px solid #cfe0fb}.tree-main strong{display:block;color:#003c9f;font-size:19px}.tree-main span{display:block;margin-top:6px;color:#4e617d;font-size:12px;line-height:1.65}.tree-main.core{background:linear-gradient(160deg,#0045bd,#006fe6)}.tree-main.core strong{color:#fff}.tree-main.core span{color:rgba(255,255,255,.87)}.tree-child{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;padding:14px}.child{padding:13px;border:1px solid #e2eaf5;border-radius:10px;background:#fbfdff}.child b{display:block;color:#0052d9;font-size:14px}.child span{display:block;margin-top:6px;color:#65748c;font-size:12px;line-height:1.6}.child strong{color:#33425c}.child em{display:inline-block;margin-top:8px;padding:2px 7px;border-radius:5px;background:#e9f1ff;color:#0052d9;font-size:11px;font-style:normal;font-weight:700}.tree-scroll{overflow-x:auto}.overall-table{width:100%;min-width:1320px;border-collapse:collapse;font-size:12px}.overall-table th{padding:10px;background:#edf4ff;color:#1d4f9d;text-align:left;border-bottom:1px solid #dce8f7;white-space:nowrap}.overall-table td{padding:10px;color:#40546f;vertical-align:top;line-height:1.6;border-bottom:1px solid #edf2f8}.overall-table tr:last-child td{border-bottom:0}.overall-table .child-name{font-weight:700;color:#0052d9;white-space:nowrap}.overall-table em{display:inline-block;padding:2px 7px;border-radius:5px;background:#e9f1ff;color:#0052d9;font-size:11px;font-style:normal;font-weight:700}.footer{text-align:center;color:#91a0b6;margin-top:30px;font-size:12px}@media(max-width:1000px){.entry-grid{grid-template-columns:repeat(2,1fr)}.tree{grid-template-columns:1fr}.tree-main{border-right:0;border-bottom:1px solid #cfe0fb}.tree-child{grid-template-columns:repeat(2,1fr)}}@media(max-width:700px){.wrap{padding:16px}.hero{padding:28px 24px}.hero h1{font-size:26px}.entry-grid,.tree-child{grid-template-columns:1fr}.section{padding:19px}}
</style>'''

SECTOR_STYLE = r'''<style>
:root{--blue:#0052d9;--ink:#172b4d;--text:#33425c;--muted:#71809a;--bg:#f4f7fb;--card:#fff;--line:#e2eaf5}*{box-sizing:border-box}body{margin:0;background:var(--bg);color:var(--ink);font-family:-apple-system,BlinkMacSystemFont,"PingFang SC","Microsoft YaHei",Arial,sans-serif}.wrap{max-width:1440px;margin:auto;padding:24px 36px 56px}.back{display:inline-block;color:#46617e;text-decoration:none;font-size:13px;margin-bottom:14px}.hero{padding:30px 34px;border-radius:21px;background:linear-gradient(135deg,#003ea6,#0052d9 60%,#1590f4);color:white}.hero h1{margin:8px 0;font-size:28px}.hero p{margin:0;color:rgba(255,255,255,.9);font-size:13px}.tabs{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin:20px 0}.tab{border:1px solid var(--line);background:#fff;border-radius:13px;padding:15px;cursor:pointer;color:#60708a;font-weight:700;font-size:15px}.tab.active{background:#0052d9;color:white;border-color:#0052d9}.view{display:none}.view.active{display:block}.section{padding:24px;background:#fff;border:1px solid var(--line);border-radius:18px;box-shadow:0 4px 16px rgba(0,82,217,.04);margin-top:14px}.section h2{margin:0;font-size:20px}.section>p{color:var(--muted);font-size:13px;line-height:1.7}.kpis{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-top:16px}.kpi{padding:17px;border:1px solid #e0ebfa;border-radius:12px;background:#f7faff}.kpi b{display:block;color:#0052d9;font-size:25px}.kpi span{display:block;margin-top:5px;color:var(--muted);font-size:12px}.insight{margin-top:16px;padding:18px;border-left:4px solid #0052d9;background:#f4f8ff;border-radius:10px;color:#3e587b;font-size:13px;line-height:1.8}.analysis-label{margin:26px 0 8px;padding-left:10px;border-left:4px solid #0052d9;color:#0045bd;font-size:15px;font-weight:700}.analysis-label:first-child{margin-top:18px}.dim-grid{display:grid;grid-template-columns:repeat(2,1fr);gap:14px;margin-top:10px}.dim{border:1px solid #e0ebfa;border-radius:14px;padding:17px}.dim.core{background:linear-gradient(145deg,#e6f0ff,#d6e8ff);border-color:#85b5ff}.dim h3{margin:0 0 10px;color:#0045bd;font-size:16px}.scroll{overflow-x:auto}.table{width:100%;min-width:620px;border-collapse:collapse;font-size:12px}.table th{background:#edf4ff;color:#164c9c;text-align:left;padding:9px;border-bottom:1px solid #dce8f7}.table td{padding:9px;color:#40546f;border-bottom:1px solid #edf2f8}.table .name{font-weight:700;color:#1f3e73}.table .hi{font-weight:700;color:#0052d9}.tree-wrap{display:flex;flex-direction:column;gap:13px}.tree{display:grid;grid-template-columns:205px minmax(0,1fr);border:1px solid var(--line);border-radius:14px;overflow:hidden}.tree-main{padding:18px;background:#e7f0ff}.tree-main.core{background:linear-gradient(160deg,#0045bd,#006fe6);color:#fff}.tree-main b{font-size:17px}.tree-main span{display:block;margin-top:5px;font-size:12px;line-height:1.6}.tree-child{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;padding:12px}.child{border:1px solid #e2eaf5;border-radius:8px;padding:11px;font-size:12px;color:#5d6d84;line-height:1.6}.child b{display:block;color:#0052d9}.tree-scroll{overflow-x:auto;padding:12px}.overall-table{width:100%;min-width:1120px;border-collapse:collapse;font-size:12px}.overall-table th{padding:9px;background:#edf4ff;color:#164c9c;text-align:left;border-bottom:1px solid #dce8f7;white-space:nowrap}.overall-table td{padding:9px;color:#40546f;vertical-align:top;line-height:1.6;border-bottom:1px solid #edf2f8}.overall-table tr:last-child td{border-bottom:0}.overall-table .child-name{font-weight:700;color:#0052d9;white-space:nowrap}.overall-table em{display:inline-block;padding:2px 7px;border-radius:5px;background:#e9f1ff;color:#0052d9;font-size:11px;font-style:normal;font-weight:700}.material-list{display:flex;flex-direction:column;gap:16px;margin-top:16px}.material{padding:18px;border:1px solid #dbe6f5;border-radius:16px;background:#fff;box-shadow:0 4px 14px rgba(0,82,217,.04)}.material-head{display:flex;justify-content:space-between;gap:16px;border-bottom:1px solid #e7eef8;padding-bottom:12px}.material-name{font-size:18px;font-weight:700}.material-sub{margin-left:7px;color:#70829d;font-size:13px}.material-kpis{display:flex;gap:13px;color:#6d7d95;font-size:12px;white-space:nowrap}.material-kpis b{color:#0052d9}.script-path{margin:12px 0 8px;color:#3a5072;font-size:13px}.learning{padding:11px 13px;background:#f0f5ff;border-left:3px solid #0052d9;border-radius:7px;color:#405a7c;font-size:12.5px;line-height:1.7}.learning b{color:#0052d9}.material-body{display:grid;grid-template-columns:150px minmax(0,1fr);gap:12px;margin-top:12px}.material-video video{display:block;width:150px;height:266px;background:#0d1828;object-fit:contain;border-radius:9px}.frames{display:grid;grid-template-columns:repeat(6,minmax(95px,1fr));gap:8px;overflow-x:auto}.frame{position:relative;min-width:95px;margin:0;overflow:hidden;border:1px solid #d9e5f5;border-radius:8px;background:#edf3fb}.frame img{width:100%;height:170px;object-fit:cover;display:block}.frame .pct{position:absolute;top:5px;left:5px;padding:2px 5px;border-radius:4px;background:#0052d9;color:white;font-size:10px;font-weight:700}.frame figcaption{text-align:center;padding:6px;color:#65758c;font-size:10px}.creative{margin-top:14px;padding:14px;background:#f7faff;border:1px solid #e0eaf8;border-radius:11px}.creative-title{color:#0052d9;font-size:13px;font-weight:700;margin-bottom:10px}.creative-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:10px 18px}.creative-group b{display:block;color:#71809a;font-size:11px;margin-bottom:5px}.creative-group span{display:inline-block;margin:0 4px 4px 0;padding:3px 7px;border:1px solid #d6e4fa;border-radius:5px;color:#2e68bd;background:#fff;font-size:11px}.creative-group.core span{background:#0052d9;color:#fff;border-color:#0052d9}.note{margin-top:12px;color:#71809a;font-size:12px;line-height:1.6}@media(max-width:1000px){.kpis,.dim-grid{grid-template-columns:1fr}.tree{grid-template-columns:1fr}.tree-child{grid-template-columns:1fr}.creative-grid{grid-template-columns:repeat(2,1fr)}.wrap{padding:16px}}@media(max-width:700px){.material-head{display:block}.material-kpis{margin-top:8px;flex-wrap:wrap}.material-body{grid-template-columns:1fr}.material-video video{width:100%;height:310px}.frames{grid-template-columns:repeat(6,100px)}.frame img{height:145px}}
</style>'''

def material_script(item):
    t = item['tags']
    category = item['category'] or '该品类'
    path = f"{t['struct']}开场呈现 {t['function']} → {t['sell']}证明 → {t['people']}场景代入 → 回应「{t['pain']}」→ CTA 收尾"
    learning = f"{category}优先把「{t['function']}」放在开场或前 3 秒，并通过{t['sell']}建立可信度；画面标签已优先采用关键帧人工核验结果。"
    return path, learning


def sector_html(payload):
    # 与箱包鞋靴标签树采用同一布局：左侧一级标签、右侧完整二级标签表。
    tree = tree_html(sector_tree(payload['id']), overall=True)
    dim_names={'struct':'制作形式','role':'出镜角色类型','scene':'拍摄场景类型','style':'穿搭风格类型','sell':'卖点表达类型','function':'功能属性','material':'材质属性','season':'适用季节','emotion':'情绪钩子类型','use':'题材/使用场景','festival':'节日/节点营销'}
    dims=[]
    for key,title in dim_names.items():
        rows=''.join(f'<tr><td class="name">{esc(x["name"])}</td><td>{"—" if x["share"] is None else str(x["share"])+"%"}</td><td class="hi">{"—" if x["ctr"] is None else str(x["ctr"])+"%"}</td><td>{"—" if x["v3"] is None else str(x["v3"])+"%"}</td><td>{"—" if x["cvr"] is None else str(x["cvr"])+"%"}</td></tr>' for x in payload['analysis'][key])
        dims.append(f'<article class="dim {"core" if key in {"struct","role","scene"} else ""}"><h3>{title}</h3><div class="scroll"><table class="table"><thead><tr><th>标签</th><th>消耗占比</th><th>CTR</th><th>3秒完播</th><th>CVR</th></tr></thead><tbody>{rows}</tbody></table></div></article>')
    # 对齐箱包鞋靴分析页：先制作形式，再画面客观要素，随后依次呈现商品、卖点、情绪维度。
    analysis_blocks = (
        '<div class="analysis-label">一、制作形式（最核心）</div><div class="dim-grid">'+dims[0]+'</div>'
        '<div class="analysis-label">二、画面客观要素</div><div class="dim-grid">'+dims[1]+dims[2]+'</div>'
        '<div class="analysis-label">三、商品属性</div><div class="dim-grid">'+dims[5]+dims[6]+dims[7]+'</div>'
        '<div class="analysis-label">四、卖点表达</div><div class="dim-grid">'+dims[4]+'</div>'
        '<div class="analysis-label">五、情绪与营销</div><div class="dim-grid">'+dims[8]+dims[3]+dims[9]+dims[10]+'</div>'
    )
    cats=''.join(f'<tr><td>{i+1}</td><td class="name">{esc(x["name"])}</td><td>{x["count"]}</td><td class="hi">{x["share"]}%</td><td class="hi">{x["ctr"]}%</td><td>{x["v3"]}%</td><td>{x["cvr"]}%</td></tr>' for i,x in enumerate(payload['categories']))
    ratios=[('2%','开场'),('20%','起势'),('40%','卖点'),('60%','演示'),('80%','信任'),('95%','收尾')]
    materials=[]
    for i,x in enumerate(payload['top']):
        t=x['tags']; frame_dir=f'../assets/frames/{payload["id"]}/{x["id"]}'
        path, learning=material_script(x)
        frames=''.join(f'<figure class="frame"><span class="pct">{pct}</span><img loading="lazy" src="{frame_dir}/{n}.jpg" onerror="this.style.display=\'none\'" alt="{label}"><figcaption>{label}</figcaption></figure>' for n,(pct,label) in enumerate(ratios,1))
        groups=[('制作形式',t['struct'],True),('出镜角色',t['role'],True),('拍摄场景',t['scene'],True),('产品卖点',t['sell'],False),('场景人群',t['people'],False),('产品功效',t['function'],False),('用户痛点',t['pain'],False),('情绪钩子',t['emotion'],False)]
        creative=''.join('<div class="creative-group'+(' core' if core else '')+'"><b>'+label+'</b><span>'+esc(value)+'</span></div>' for label,value,core in groups)
        materials.append(f"""<article class="material"><div class="material-head"><div class="material-name">#{i+1} {esc(x['category'] or '未标注类目')}<span class="material-sub">{payload['emoji']} {payload['name']}</span></div><div class="material-kpis"><span>CTR <b>{x['ctr']:.2f}%</b></span><span>3s <b>{x['v3']:.1f}%</b></span><span>CVR <b>{x['cvr']:.2f}%</b></span></div></div><div class="script-path"><b>叙事路径：</b>{esc(path)}</div><div class="learning"><b>Learning：</b>{esc(learning)}</div><div class="material-body"><div class="material-video"><video controls muted preload="metadata" src="{esc(x['url'])}"></video></div><div class="frames">{frames}</div></div><div class="creative"><div class="creative-title">🏷 创意标签</div><div class="creative-grid">{creative}</div></div></article>""")
    return f"""<!doctype html><html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>{payload['name']}创意分析</title>{SECTOR_STYLE}</head><body><main class="wrap"><a class="back" href="../">← 返回服饰行业总站</a><header class="hero"><div>AI 创意洞察 · {payload['name']}赛道</div><h1>{payload['emoji']} {payload['name']}创意分析</h1><p>{payload['desc']} · 标签树、效果分析与 Top 素材逐帧创意拆解。</p></header><div class="tabs"><button class="tab active" data-tab="tree">🌳 素材标签树结构</button><button class="tab" data-tab="analysis">🏷 标签分析和结论</button><button class="tab" data-tab="materials">🎬 素材内容 · 逐帧剖析</button></div><div class="view active" id="tree"><section class="section"><h2>{payload['name']}赛道标签树结构</h2><p>内容形式、画面客观要素使用统一核心标签；商品属性、卖点表达、情绪营销已按 {payload['name']} 的实际类目与素材特点调整枚举值。</p><div class="tree-wrap">{tree}</div></section></div><div class="view" id="analysis"><section class="section"><h2>大盘标签分析 · 一分钟看懂本期</h2><div class="kpis"><div class="kpi"><b>{payload['metrics']['materials']}</b><span>有效去重素材</span></div><div class="kpi"><b>{payload['metrics']['ctr']}%</b><span>消耗加权 CTR</span></div><div class="kpi"><b>{payload['metrics']['v3']}%</b><span>消耗加权 3秒完播</span></div><div class="kpi"><b>{payload['metrics']['cvr']}%</b><span>消耗加权 浅层CVR</span></div></div><div class="insight">{esc(payload['narrative'])}</div></section><section class="section"><h2>各标签维度明细</h2><p>严格按箱包鞋靴分析页的逻辑展示：制作形式优先，其次是画面客观要素，再下钻商品属性、卖点表达和情绪营销。</p>{analysis_blocks}</section><section class="section"><h2>细分类目效果</h2><div class="scroll"><table class="table"><thead><tr><th>#</th><th>细分类目</th><th>素材数</th><th>消耗占比</th><th>CTR</th><th>3秒完播</th><th>CVR</th></tr></thead><tbody>{cats}</tbody></table></div></section></div><div class="view" id="materials"><section class="section"><h2>Top 跑量素材逐帧剖析（最多 12 条）</h2><p>按消耗排序；每条均展示视频、6 个关键帧、叙事路径、Learning 与创意标签，布局与箱包鞋靴深度站保持一致。</p><div class="material-list">{''.join(materials)}</div><div class="note">{esc(payload['note'])}</div></section></div></main><script>document.querySelectorAll('.tab').forEach(b=>b.onclick=()=>{{document.querySelectorAll('.tab').forEach(x=>x.classList.toggle('active',x===b));document.querySelectorAll('.view').forEach(x=>x.classList.toggle('active',x.id===b.dataset.tab))}})</script></body></html>"""


def main():
    if not XLSX.exists(): raise FileNotFoundError(XLSX)
    DOCS.mkdir(parents=True, exist_ok=True); DATA_DIR.mkdir(parents=True, exist_ok=True)
    # 首页不加载行业数据；行业页各自独立构建。
    (DOCS/'index.html').write_text(home_html(), encoding='utf-8')
    all_data={}
    for key,meta in SECTORS.items():
        payload=sector_payload(key); all_data[key]=payload
        target=DOCS/key; target.mkdir(parents=True, exist_ok=True)
        (target/'index.html').write_text(sector_html(payload),encoding='utf-8')
    (DATA_DIR/'sector_snapshot.json').write_text(json.dumps(all_data,ensure_ascii=False,indent=2),encoding='utf-8')
    # 同步已抽取关键帧到 Pages 静态目录；未完成的帧会在下一次构建自动补入。
    source_frames = FRAME_ROOT
    target_frames = DOCS / 'assets' / 'frames'
    if source_frames.exists():
        target_frames.parent.mkdir(parents=True, exist_ok=True)
        shutil.copytree(source_frames, target_frames, dirs_exist_ok=True)
    print('[✓] 总站：行业入口 + 服饰整体标签树')
    for key,x in all_data.items(): print(f"[✓] {x['name']}深度页：{x['metrics']['materials']} 条素材")

if __name__ == '__main__': main()
