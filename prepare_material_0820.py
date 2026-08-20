#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""素材8.20 数据准备：分三类 + 抽取待打标素材前15秒关键帧。"""
import json, hashlib, os, statistics
from collections import defaultdict
from openpyxl import load_workbook

XLSX = '/Users/chenxiaomei/Desktop/素材8.20.xlsx'
BASE = '/Users/chenxiaomei/CodeBuddy/20260508175441/apparel-creative-analysis'
DATA = os.path.join(BASE, 'data')
os.makedirs(DATA, exist_ok=True)


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def load():
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb['Sheet1']
    raw = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        t, url, pid, pname, spend, ctr, cvr, dur, v3 = r
        if not isinstance(url, str) or not url.startswith('http'):
            continue
        raw.append({'date': str(t)[:10].replace('/', '-'), 'url': url, 'pid': str(pid or ''),
                    'pname': str(pname or '').strip(), 'spend': num(spend), 'ctr': num(ctr),
                    'cvr': num(cvr), 'dur': num(dur), 'v3': num(v3)})
    return raw


def aggregate(raw):
    """按 url 聚合多天数据，并计算 ctr 趋势斜率。"""
    g = defaultdict(list)
    for x in raw:
        g[x['url']].append(x)
    out = []
    for url, days in g.items():
        days.sort(key=lambda d: d['date'])
        tw = sum(d['spend'] for d in days)
        w = lambda k: (sum(d['spend'] * d[k] for d in days) / tw) if tw else statistics.fmean(d[k] for d in days)
        # ctr 趋势：后半段均值 - 前半段均值（仅多天素材有效）
        slope = 0.0
        if len(days) >= 2:
            half = max(len(days) // 2, 1)
            first = statistics.fmean(d['ctr'] for d in days[:half])
            last = statistics.fmean(d['ctr'] for d in days[half:])
            slope = last - first
        out.append({'url': url, 'id': hashlib.md5(url.encode()).hexdigest()[:10],
                    'pid': days[-1]['pid'], 'pname': days[-1]['pname'],
                    'spend': round(tw, 4), 'ctr': round(w('ctr'), 3), 'cvr': round(w('cvr'), 3),
                    'dur': round(w('dur'), 2), 'v3': round(w('v3'), 2),
                    'days': len(days), 'slope': round(slope, 3),
                    'first_date': days[0]['date'], 'last_date': days[-1]['date']})
    return out


def classify(items):
    """好：ctr 高为主，播放时长/3秒完播为次要；差：ctr 低；上升：ctr 斜率为正。"""
    ctrs = sorted(x['ctr'] for x in items)
    n = len(ctrs)
    p = lambda q: ctrs[min(int(n * q), n - 1)]
    med_dur = statistics.median(x['dur'] for x in items)
    med_v3 = statistics.median(x['v3'] for x in items)

    def score(x):
        # ctr 为主(权重1.0)，播放时长与3秒完播为次要参考(各0.25)
        return x['ctr'] + 0.25 * (x['dur'] / med_dur if med_dur else 0) + 0.25 * (x['v3'] / med_v3 if med_v3 else 0)

    good = sorted(items, key=score, reverse=True)[:200]
    bad = sorted(items, key=lambda x: (x['ctr'], x['v3']))[:200]
    rising_pool = [x for x in items if x['days'] >= 2 and x['slope'] > 0]
    rising = sorted(rising_pool, key=lambda x: x['slope'], reverse=True)[:100]
    return good, bad, rising, {'ctr_p25': p(0.25), 'ctr_p50': p(0.5), 'ctr_p75': p(0.75),
                               'med_dur': round(med_dur, 2), 'med_v3': round(med_v3, 2)}


if __name__ == '__main__':
    raw = load()
    items = aggregate(raw)
    good, bad, rising, stat = classify(items)
    by_spend = sorted(items, key=lambda x: x['spend'], reverse=True)
    demo30 = by_spend[:30]                       # Tab2 打标示范
    golden50 = sorted(good, key=lambda x: x['spend'], reverse=True)[:50]  # Tab4 黄金公式
    need = {x['id']: x for x in demo30 + golden50}
    json.dump({'stat': stat, 'total': len(items), 'items': items, 'good': good, 'bad': bad,
               'rising': rising, 'demo30': demo30, 'golden50': golden50,
               'need_frames': list(need.values())},
              open(os.path.join(DATA, 'material_0820.json'), 'w', encoding='utf-8'),
              ensure_ascii=False, indent=1)
    print('去重素材', len(items), '| 好', len(good), '| 差', len(bad), '| 上升', len(rising))
    print('需抽帧素材', len(need))
    print('统计', stat)
    print('Top1 消耗', by_spend[0]['pname'], by_spend[0]['spend'])
